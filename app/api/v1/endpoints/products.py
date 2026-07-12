import csv
from io import BytesIO, StringIO
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import logging
from pydantic import ValidationError

from app.core.database import get_db
from app.models.product import Product
from app.models.product_code import ProductCode
from app.schemas.product import ProductCreate, ProductUpdate, ProductResponse, BulkProductResponse
from app.schemas.error import ErrorResponse

router = APIRouter()
logger = logging.getLogger(__name__)

_ERROR_RESPONSES = {
    400: {"model": ErrorResponse, "description": "Bad request"},
    404: {"model": ErrorResponse, "description": "Not found"},
    422: {"model": ErrorResponse, "description": "Validation error"},
}

_UPLOAD_REQUIRED_COLUMNS = {"product_code_id", "product_name"}
_UPLOAD_ALLOWED_COLUMNS = {"product_code_id", "product_name", "brand", "category", "ai_code", "type"}


def _clean_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed if trimmed else None
    return value


def _normalize_row(raw_row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in raw_row.items():
        if key is None:
            continue
        clean_key = str(key).strip().lower()
        if not clean_key:
            continue
        normalized[clean_key] = _clean_cell_value(value)
    return normalized


def _extract_rows_from_csv(file_content: bytes) -> list[tuple[int, dict[str, Any]]]:
    try:
        text = file_content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="CSV file must be UTF-8 encoded") from exc

    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV file is missing header row")

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(reader, start=2):
        normalized = _normalize_row(row)
        if not normalized or all(v is None for v in normalized.values()):
            continue
        rows.append((row_number, normalized))
    return rows


def _extract_rows_from_excel(file_content: bytes) -> list[tuple[int, dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel support requires openpyxl") from exc

    try:
        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid Excel file") from exc

    sheet = workbook.active
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        raise HTTPException(status_code=400, detail="Excel file is missing header row")

    headers = [str(cell).strip() if cell is not None else "" for cell in header_cells]
    if not any(headers):
        raise HTTPException(status_code=400, detail="Excel file header row is empty")

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        normalized = _normalize_row(raw_row)
        if not normalized or all(v is None for v in normalized.values()):
            continue
        rows.append((row_number, normalized))
    return rows


def _parse_upload_rows(filename: str, file_content: bytes) -> list[tuple[int, dict[str, Any]]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _extract_rows_from_csv(file_content)
    if lower_name.endswith(".xlsx"):
        return _extract_rows_from_excel(file_content)
    raise HTTPException(status_code=400, detail="Unsupported file type. Use .csv or .xlsx")


def _validate_upload_columns(rows: list[tuple[int, dict[str, Any]]]) -> None:
    if not rows:
        return
    columns = set().union(*(row.keys() for _, row in rows))
    missing = sorted(_UPLOAD_REQUIRED_COLUMNS - columns)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    unexpected = sorted(columns - _UPLOAD_ALLOWED_COLUMNS)
    if unexpected:
        raise HTTPException(status_code=400, detail=f"Unsupported columns: {', '.join(unexpected)}")


def _build_csv_template() -> bytes:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=["product_code_id", "product_name", "brand", "category", "ai_code", "type"])
    writer.writeheader()
    writer.writerow(
        {
            "product_code_id": 1,
            "product_name": "Sample Product A",
            "brand": "Sample Brand",
            "category": "Beverages",
            "ai_code": "AI-001",
            "type": "own",
        }
    )
    writer.writerow(
        {
            "product_code_id": 2,
            "product_name": "Sample Product B",
            "brand": "Comp Brand",
            "category": "Snacks",
            "ai_code": "AI-002",
            "type": "competitor",
        }
    )
    return output.getvalue().encode("utf-8")


def _build_xlsx_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel support requires openpyxl") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products_template"
    sheet.append(["product_code_id", "product_name", "brand", "category", "ai_code", "type"])
    sheet.append([1, "Sample Product A", "Sample Brand", "Beverages", "AI-001", "own"])
    sheet.append([2, "Sample Product B", "Comp Brand", "Snacks", "AI-002", "competitor"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ─────────────────────────────────────────────
# CREATE
# ─────────────────────────────────────────────

@router.post("/", response_model=ProductResponse, responses={400: _ERROR_RESPONSES[400], 422: _ERROR_RESPONSES[422]}, summary="Create a product")
def create_product(product: ProductCreate, db: Session = Depends(get_db)):
    product_code = db.query(ProductCode).filter(ProductCode.id == product.product_code_id).first()
    if not product_code:
        raise HTTPException(status_code=400, detail="Invalid product_code_id")

    existing = db.query(Product).filter(
        Product.product_name == product.product_name
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail="Product already exists")

    db_product = Product(**product.dict())
    db.add(db_product)
    db.commit()
    db.refresh(db_product)

    logger.info(f"Product created: {db_product.product_name}")
    return db_product


@router.post("/bulk", response_model=BulkProductResponse, responses={422: _ERROR_RESPONSES[422]}, summary="Create multiple products at once")
def create_products_bulk(products: list[ProductCreate], db: Session = Depends(get_db)):
    created = []
    skipped = []
    
    try:
        # First pass: validate all items (check for existing duplicates)
        for item in products:
            product_code = db.query(ProductCode).filter(ProductCode.id == item.product_code_id).first()
            if not product_code:
                raise HTTPException(status_code=400, detail=f"Invalid product_code_id for {item.product_name}")
            existing = db.query(Product).filter(Product.product_name == item.product_name).first()
            if existing:
                skipped.append(item.product_name)
        
        # Second pass: add only non-duplicates to session (not yet committed)
        for item in products:
            if item.product_name not in skipped:
                obj = Product(**item.dict())
                db.add(obj)
                created.append(obj)
        
        # Commit atomically - all products commit together or none at all
        db.commit()
        
        # Refresh objects after successful commit
        for obj in created:
            db.refresh(obj)
        
        logger.info(f"Bulk create atomic: {len(created)} created, {len(skipped)} skipped")
        return {"created": created, "skipped": skipped}
    
    except Exception as e:
        # Rollback entire transaction if anything fails
        db.rollback()
        logger.exception(f"Bulk insert failed, rolled back all changes: {e}")
        raise HTTPException(status_code=500, detail=f"Bulk insert failed and rolled back: {str(e)}")


@router.post(
    "/bulk/upload",
    response_model=BulkProductResponse,
    responses={400: _ERROR_RESPONSES[400], 422: _ERROR_RESPONSES[422]},
    summary="Bulk upload products from CSV or Excel",
)
async def create_products_bulk_upload(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    file_content = await file.read()
    if not file_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    parsed_rows = _parse_upload_rows(file.filename, file_content)
    if not parsed_rows:
        raise HTTPException(status_code=400, detail="No data rows found in the uploaded file")

    _validate_upload_columns(parsed_rows)

    validation_errors: list[str] = []
    validated_payloads: list[ProductCreate] = []
    for row_number, row in parsed_rows:
        filtered_row = {k: v for k, v in row.items() if k in _UPLOAD_ALLOWED_COLUMNS}
        try:
            payload = ProductCreate(**filtered_row)
            validated_payloads.append(payload)
        except ValidationError as exc:
            validation_errors.append(f"Row {row_number}: {exc.errors()}")

    if validation_errors:
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "errors": validation_errors})

    code_ids = {payload.product_code_id for payload in validated_payloads}
    valid_code_ids = {
        code_id
        for (code_id,) in db.query(ProductCode.id).filter(ProductCode.id.in_(code_ids)).all()
    }
    missing_code_ids = sorted(code_ids - valid_code_ids)
    if missing_code_ids:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid product_code_id values: {', '.join(map(str, missing_code_ids))}",
        )

    seen_names: set[str] = set()
    duplicate_names_in_file: set[str] = set()
    for payload in validated_payloads:
        if payload.product_name in seen_names:
            duplicate_names_in_file.add(payload.product_name)
        else:
            seen_names.add(payload.product_name)

    existing_names = {
        name
        for (name,) in db.query(Product.product_name).filter(Product.product_name.in_(seen_names)).all()
    }

    skipped_names = set(duplicate_names_in_file) | existing_names
    created: list[Product] = []

    try:
        for payload in validated_payloads:
            if payload.product_name in skipped_names:
                continue

            obj = Product(**payload.model_dump())
            db.add(obj)
            created.append(obj)

        db.commit()

        for obj in created:
            db.refresh(obj)

        logger.info(f"Bulk upload complete: {len(created)} created, {len(skipped_names)} skipped")
        return {"created": created, "skipped": sorted(skipped_names)}
    except Exception as exc:
        db.rollback()
        logger.exception(f"Bulk upload failed, transaction rolled back: {exc}")
        raise HTTPException(status_code=500, detail="Bulk upload failed and was rolled back")


@router.get(
    "/bulk/template",
    summary="Download bulk upload template for products",
    responses={400: _ERROR_RESPONSES[400]},
)
def download_products_bulk_template(
    format: str = Query("csv", description="Template format: csv or xlsx"),
):
    template_format = format.strip().lower()

    if template_format == "csv":
        content = _build_csv_template()
        filename = "products_bulk_template.csv"
        media_type = "text/csv"
    elif template_format == "xlsx":
        content = _build_xlsx_template()
        filename = "products_bulk_template.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'xlsx'")

    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(BytesIO(content), media_type=media_type, headers=headers)


# ─────────────────────────────────────────────
# READ
# ─────────────────────────────────────────────

@router.get("/", response_model=list[ProductResponse], responses={422: _ERROR_RESPONSES[422]}, summary="List all products")
def get_all_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db)
):
    return db.query(Product).offset(skip).limit(limit).all()


@router.get("/search/", response_model=list[ProductResponse], responses={422: _ERROR_RESPONSES[422]}, summary="Search products by name, brand, category or type")
def search_products(
    product_code_id: int = Query(None, description="Filter by product code ID"),
    name: str = Query(None, description="Partial product name"),
    brand: str = Query(None, description="Partial brand name"),
    category: str = Query(None, description="Partial category"),
    type: str = Query(None, description="own / competitor"),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db)
):
    q = db.query(Product)
    if product_code_id is not None:
        q = q.filter(Product.product_code_id == product_code_id)
    if name:
        q = q.filter(Product.product_name.ilike(f"%{name}%"))
    if brand:
        q = q.filter(Product.brand.ilike(f"%{brand}%"))
    if category:
        q = q.filter(Product.category.ilike(f"%{category}%"))
    if type:
        q = q.filter(Product.type == type)
    return q.offset(skip).limit(limit).all()


@router.get("/by-name/{product_name}", response_model=ProductResponse, responses={404: _ERROR_RESPONSES[404]}, summary="Get product by name")
def get_product_by_name(product_name: str, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.product_name == product_name).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.get("/{product_id}", response_model=ProductResponse, responses={404: _ERROR_RESPONSES[404]}, summary="Get product by ID")
def get_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()

    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    return product


# ─────────────────────────────────────────────
# UPDATE
# ─────────────────────────────────────────────

@router.put("/by-name/{product_name}", response_model=ProductResponse, responses={400: _ERROR_RESPONSES[400], 404: _ERROR_RESPONSES[404], 422: _ERROR_RESPONSES[422]}, summary="Update product by name")
def update_product_by_name(
    product_name: str,
    product_update: ProductUpdate,
    db: Session = Depends(get_db)
):
    product = db.query(Product).filter(Product.product_name == product_name).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    if product_update.product_name and product_update.product_name != product_name:
        exists = db.query(Product).filter(Product.product_name == product_update.product_name).first()
        if exists:
            raise HTTPException(status_code=400, detail="Product name already exists")

    if product_update.product_code_id is not None:
        product_code = db.query(ProductCode).filter(ProductCode.id == product_update.product_code_id).first()
        if not product_code:
            raise HTTPException(status_code=400, detail="Invalid product_code_id")

    for key, value in product_update.dict(exclude_unset=True).items():
        setattr(product, key, value)

    db.commit()
    db.refresh(product)

    logger.info(f"Product updated by name: {product_name}")
    return product


@router.put("/{product_id}", response_model=ProductResponse, responses={400: _ERROR_RESPONSES[400], 404: _ERROR_RESPONSES[404], 422: _ERROR_RESPONSES[422]}, summary="Update product by ID")
def update_product(
    product_id: int,
    product_update: ProductUpdate,
    db: Session = Depends(get_db)
):
    product = db.query(Product).filter(Product.id == product_id).first()

    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    if product_update.product_name:
        exists = db.query(Product).filter(
            Product.product_name == product_update.product_name,
            Product.id != product_id
        ).first()
        if exists:
            raise HTTPException(status_code=400, detail="Product name already exists")

    if product_update.product_code_id is not None:
        product_code = db.query(ProductCode).filter(ProductCode.id == product_update.product_code_id).first()
        if not product_code:
            raise HTTPException(status_code=400, detail="Invalid product_code_id")

    for key, value in product_update.dict(exclude_unset=True).items():
        setattr(product, key, value)

    db.commit()
    db.refresh(product)

    logger.info(f"Product updated: {product_id}")
    return product


# ─────────────────────────────────────────────
# DELETE
# ─────────────────────────────────────────────

@router.delete("/by-name/{product_name}", responses={404: _ERROR_RESPONSES[404]}, summary="Delete product by name")
def delete_product_by_name(product_name: str, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.product_name == product_name).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(product)
    db.commit()

    logger.info(f"Product deleted by name: {product_name}")
    return {"message": f"Product '{product_name}' deleted successfully"}


@router.delete("/{product_id}", responses={404: _ERROR_RESPONSES[404]}, summary="Delete product by ID")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()

    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(product)
    db.commit()

    logger.info(f"Product deleted: {product_id}")
    return {"message": "Product deleted successfully"}

